# -*- coding: utf-8 -*-
"""
Created on Thu Aug 18 10:30:38 2022

@author: wytheY
"""
from iFinDPy import *
#from tick_trade_api import DatafeedHqGenerator
#from tick_trade_api import MindgoHqGenerator
import pandas as pd
import openpyxl as op
import numpy as np
import dateutil.relativedelta
import time
import datetime
import csv
from openpyxl import load_workbook
import calendar
import os
global null
import re
null=''  

#THS_iFinDLogin('ztzgsx002','609226')
today_str = datetime.datetime.now()

wb = load_workbook(filename=r'C:\Users\wytheY\.spyder-py3\pandas 運用\name.xlsx')
ws = wb.get_sheet_by_name('Sheet1') 
#writer = pd.ExcelWriter('new.xlsx')#这里是创建了可写入不同sheet的文件?
#指定当前工作表名称,注意表名的大小写
#print('工作表列数：',ws.max_column)
#print('工作表行数：',ws.max_row)
with open("report.csv",'w+',newline='') as t2:
    title=["DAY","ID","DATA","Maincontract","Submaincontract"]
    writer2=csv.writer(t2)#这一步是创建一个csv的写入器
    writer2.writerow(title)#写入标签
t2.close
with open("results.csv",'w+',newline='') as t3:
    title=["日期","做空","做多","主力空","主力多"]
    writer3=csv.writer(t3)#这一步是创建一个csv的写入器
    writer3.writerow(title)#写入标签
    with open("S_DQ_OI.csv",'w+',newline='') as t1:#numline是来控制空的行数的
        tit=["OI",]
        writer=csv.writer(t1)#这一步是创建一个csv的写入器
        writer.writerow(tit)#写入标签
        for p in range(1,5):    
             for k in range(1, calendar.mdays[p]+1): 
               shift_date = p,'月',k,'日',
               shift_date = str(shift_date)
               shift_date = shift_date.replace("'","")
               shift_date = shift_date.replace(",","")
               shift_date = shift_date.replace("(","")
               shift_date = shift_date.replace(")","")
               shift_date = shift_date.replace(" ","")
               print(p,'月',k,'日',)
               df = pd.read_csv("历史数据2.csv",)
               timechange3 = str(p).rjust(2,'0')
               timechange4 = str(k).rjust(2,'0')
               timechange1 = "2015"+timechange3+timechange4+"230000"
               #print(timechange1)  # <class 'int'> 
               timechange2 = str(timechange1)
               #print(timechange2)  # <class 'str'>
               from datetime import datetime
               shift_datetime = datetime.strptime(timechange2, '%Y%m%d%H%M%S')  # 字符串 -> 时间
               shift_date1 = shift_datetime.strftime("%Y-%m-%d %H:%M:%S")
               shift_date3 = shift_date1[0:10]
           #   print(shift_date1)
           #   timechange3 = str(p).rjust(2,'0')
           #   timechange4 = str(k).rjust(2,'0')
               timechange1 = "2015"+timechange3+timechange4+"235959"
               #print(timechange1)  # <class 'int'> 
           #   timechange2 = str(timechange1)
               timechange2 = timechange1[0:8]
               timechange2 = int(timechange2)
           #   print(timechange2)
               #print(timechange2)  # <class 'str'>
               df = df[df['TRADE_DT'].isin([timechange2])] #找出数值对应行#df = pd.read_csv("RBdata.csv",)
               df4 = pd.read_csv('S_INFO_WINDCODE2.csv',header=0)
    #          df4 = df4.to_csv('S_INFO_WINDCODE2.csv', header=False)
               k = 0
               arr = list
        
        
               if df.empty:
                   print("空日")
               else:
                   with open("report.csv",'a+',newline='') as t2:
                      #title=["DAY","ID","DATA","Maincontract","Submaincontract"]
                       writer2=csv.writer(t2)#这一步是创建一个csv的写入器
                    #  if t2.tell() <= 10:
                    #     writer2.writerow(title)#写入标签
                #      print(df)
                       for line2 in df4.WINDCODE:#遍历列表中的元素   
                         if len(line2) == 1 :
                            print(line2)
                            df = pd.read_csv("历史数据2.csv",)
                            df = df[df['TRADE_DT'].isin([timechange2])]
                            df2 = df.S_INFO_WINDCODE                 
                            for line in df2:
                               if line2 in line:
                                  df3 = df[df['S_INFO_WINDCODE'].isin([line])]
                #                 sum(df3.S_DQ_OI.values)
                                  s3x = str(df3.S_INFO_WINDCODE.values)
                                  s3 = s3x[2:4]
                                  s4 = s3[1:2]                                 
                                  s4 = s4.isdigit()
                                  if s4 == True:
                                     df4 = df[df['S_INFO_WINDCODE'].isin([s3x])]
                                     arr = sum(df3.S_DQ_OI.values)
                                     arr = str(arr)
                    #                data="This is a string"#从excel读的数据
                    #                temp=[arr]
                    #                abnormal_data.append(temp)
                                     writer.writerow([arr])#写入标签
                   #                 print(arr)
                   #                 t1.close()
                   #                 df = pd.read_csv("S_DQ_OI.csv")
                   #                 print(df.OI)     
                   #                 re.sub("[^A-Za-z]","",s)
                         else:
                             print(line2)
            #                if line2 != "WINDCODE":   
                             df = pd.read_csv("历史数据2.csv",)
                             df = df[df['TRADE_DT'].isin([timechange2])]
                             df2 = df.S_INFO_WINDCODE                 
                             for line in df2:
                                if line2 in line:
                                   df3 = df[df['S_INFO_WINDCODE'].isin([line])]
                #                  sum(df3.S_DQ_OI.values)
                                   arr = sum(df3.S_DQ_OI.values)
                                   arr = str(arr)
                #                  data="This is a string"#从excel读的数据
                #                  temp=[arr]
                #                  abnormal_data.append(temp)
                                   writer.writerow([arr])#写入标签
                #                  print(arr)
                #                  t1.close()
                #                  df = pd.read_csv("S_DQ_OI.csv")
                #                  print(df.OI)     
                         t1.seek(2)
                         f = t1.read()
                         
                         print(f)
                         list1 = f
                         list2 = list1.split("\r\n") #or list2 = str2.split(" ")
                      #  print (list2)
                         while "" in list2:
                             list2.remove("")
                      #  print (list2)
                         list2 = list(map(float,list2))
                      #  print (list2)
                      #  list3 = [x for x in list2 if np.isnan(x) == False]#去除nan   
                         arr1 = list2                 
                         def bubbleSort(arr1):
                                n = len(arr1)
                                # 遍历所有数组元素
                                for i in range(n):
                                    # Last i elements are already in place
                                        for j in range(0, n-i-1):
                                            if arr1[j] > arr1[j+1] :
                                               arr1[j], arr1[j+1] = arr1[j+1], arr1[j]
                         bubbleSort(arr1)
                            
                           #print ("排序后的数组:")
                         for i in range(len(arr1)):
                                #print (arr[i])
                                openInterest1 = max(arr1)
                                openInterest2 = arr1[i-1]
                         """
                         for line3 in df.S_INFO_WINDCODE:
                             if 
                         """
                         
                         #print (openInterest1)
                         #print (openInterest2)
                         t1.seek(2)
                         t1.truncate()   #清空文件t1.truncate()
                           
                         df = pd.read_csv("历史数据2.csv",)
                         df = df[df['TRADE_DT'].isin([timechange2])]
                         df = df.loc[df['S_INFO_WINDCODE'].str.contains(line2)] #筛选（包含）
                         df = df[df['S_DQ_OI'].isin([openInterest1])]
                         df = df.head(1) #并非完美,需要上部加入一个for语句才能够完美
                         OI1 = df.S_INFO_WINDCODE.values
                         if df.empty:
                               print("当日",line2,"无交易")
                         else:
                             print(df.S_DQ_CLOSE.values)
                             c = int(df.S_DQ_CLOSE.values)
                            #print(df)
                             df = pd.read_csv("历史数据2.csv",)
                             df = df[df['TRADE_DT'].isin([timechange2])]   
                             df = df.loc[df['S_INFO_WINDCODE'].str.contains(line2)] #筛选（包含）
                             df = df[df['S_DQ_OI'].isin([openInterest2])]
                             df = df.head(1)
                             OI2 = df.S_INFO_WINDCODE.values
                             print(df.S_DQ_CLOSE.values)
                             """
                             string = str(df.S_DQ_CLOSE)
                             x = 0
                             for i in string:
                                 if i == ".":
                                     x = x+1
                             if x >= 2:
                                 list3 = str(df.S_DQ_CLOSE).split(" ") #or list2 = str2.split(" ")
                                 print(list3)
                                 num_list_new = []   # 新建空列表，用以存储提取的数值
                                 a = ''   # 将空值赋值给a
                                 for i in list3:    # 将字符串进行遍历
                                     if str.isdigit(i):    # 判断i是否为数字，如果“是”返回True，“不是”返回False
                                            a += i   # 如果i是数字格式，将i以字符串格式加到a上
                                     else:
                                            a += " "  # 如果i不是数字格式，将“ ”（空格）加到a上
                                    # 遍历后，a的值为：
                                    #       198   4747    12305        15498915                          105    386379            #177  4217    14645390        21  530    853525  
                                    #数字与数字之间存在许多空格，所以需要对字符串a按''进行分割。
                                 num_list = a.split(" ")  # 按''进行分割，此时a由字符串格式编程列表
                                 print("num_list is \n", num_list)  
                                    # 分割后，a的值为：['', '', '', '', '', '', '198', '', '', '4747', '', '', '', '12305', '', '', '', '', '', '', '', '15498915', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '105', '', '', '', '386379', '', '', '', '', '', '', '', '', '', '', '', '177', '', '4217', '', '', '', '14645390', '', '', '', '', '', '', '', '21', '', '530', '', '', '', '853525', '', '', '', '', '', '']。应该去掉空格，并将字符串格式的数字转化为int格式（主要是字符串中的数字皆为整数，如果有小数，应转化为float格式）。
                                 for i in num_list:  # 对列表a，进行遍历
                                     try:  # try 结构体，防止出错后直接退出程序
                                         if int(i) > 0:
                                                num_list_new.append(int(i))  # 如果列表a的元素为数字，则赋值给num_list_new
                                         else:
                                                pass     # 如果不是数字，pass
                                     except:
                                            pass
                                 print("num_list is \n", num_list_new)
            
                                 bubbleSort(num_list_new)
                                 i = 0
                                 for i in range(len(list3)):
                                        #print (arr[i])
                                        openInterest1 = max(arr1)
                                        openInterest2 = arr1[i-1]
                             
                             else:    
                             """
            
                             b = int(df.S_DQ_CLOSE.values)                 
                            #print(df) #目前有误
                             OI1 = str(OI1)
                             OI2 = str(OI2)
                                                         
                             num_list_new1 = []   # 新建空列表，用以存储提取的数值
                             a = ' '   # 将空值赋值给a
                             for i in OI1:    # 将字符串进行遍历
                                    if str.isdigit(i):    # 判断i是否为数字，如果“是”返回True，“不是”返回False
                                        a += i   # 如果i是数字格式，将i以字符串格式加到a上
                                    else:
                                        a += " "  # 如果i不是数字格式，将“ ”（空格）加到a上
                                # 遍历后，a的值为：
                                #数字与数字之间存在许多空格，所以需要对字符串a按''进行分割。
                             num_list = a.split(" ")  # 按''进行分割，此时a由字符串格式编程列表
                             # print("num_list is \n", num_list)  
                                # 分割后，a的值为：['', '', '', '', '', '', '198', '', '', '4747', '', '', '', '12305', '', '', '', '', '', '', '', '15498915', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '105', '', '', '', '386379', '', '', '', '', '', '', '', '', '', '', '', '177', '', '4217', '', '', '', '14645390', '', '', '', '', '', '', '', '21', '', '530', '', '', '', '853525', '', '', '', '', '', '']。应该去掉空格，并将字符串格式的数字转化为int格式（主要是字符串中的数字皆为整数，如果有小数，应转化为float格式）。
                             for i in num_list:  # 对列表a，进行遍历
                                    try:  # try 结构体，防止出错后直接退出程序
                                        if int(i) > 0:
                                            num_list_new1.append(int(i))  # 如果列表a的元素为数字，则赋值给num_list_new
                                        else:
                                            pass     # 如果不是数字，pass
                                    except:
                                        pass
                             # print(num_list_new1)                
                                
                             num_list_new2 = []   # 新建空列表，用以存储提取的数值
                             a = ''   # 将空值赋值给a
                             for i in OI2:    # 将字符串进行遍历
                                    if str.isdigit(i):    # 判断i是否为数字，如果“是”返回True，“不是”返回False
                                        a += i   # 如果i是数字格式，将i以字符串格式加到a上
                                    else:
                                        a += " "  # 如果i不是数字格式，将“ ”（空格）加到a上
                                # 遍历后，a的值为：
                                #数字与数字之间存在许多空格，所以需要对字符串a按''进行分割。
                             num_list = a.split(" ")  # 按''进行分割，此时a由字符串格式编程列表
                             # print("num_list is \n", num_list)  
                                # 分割后，a的值为：['', '', '', '', '', '', '198', '', '', '4747', '', '', '', '12305', '', '', '', '', '', '', '', '15498915', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '105', '', '', '', '386379', '', '', '', '', '', '', '', '', '', '', '', '177', '', '4217', '', '', '', '14645390', '', '', '', '', '', '', '', '21', '', '530', '', '', '', '853525', '', '', '', '', '', '']。应该去掉空格，并将字符串格式的数字转化为int格式（主要是字符串中的数字皆为整数，如果有小数，应转化为float格式）。
                             for i in num_list:  # 对列表a，进行遍历
                                    try:  # try 结构体，防止出错后直接退出程序
                                        if int(i) > 0:
                                            num_list_new2.append(int(i))  # 如果列表a的元素为数字，则赋值给num_list_new
                                        else:
                                            pass     # 如果不是数字，pass
                                    except:
                                        pass
                              #print(num_list_new2)
                               #https://blog.csdn.net/weixin_44816589/article/details/103917063                
                               
                             numN1 = num_list_new1
                             #numN1 = [int(i) for i in numN1]
                             numN1 = str(numN1)
                             numN1 = numN1.replace("[","")
                             numN1 = numN1.replace("]","")
                             t_str1 = '20'+numN1
                             """   
                               #取出日期大小
                               time_filter = filter(str.isdigit, thscode2)
                   
                               time_list = list(time_filter)       # ['2', '0', '1', '9', '0', '9', '0', '4', '1', '1', '0', '0']
                               time_str = "".join(time_list)       # 转为str    201909041100
                             """
                             numN2 = num_list_new2
                             #numN2 = [int(i) for i in numN2]     # num_list为x迭代的list,lambda输入x,输出int(x)
                             numN2 = str(numN2)
                             numN2 = numN2.replace("[","")
                             numN2 = numN2.replace("]","")
                             t_str2 = '20'+numN2
                   
                               #在datetime模块中有timedelta类，这个类的对象用于表示一个时间间隔，比如两个日#期或者时间的差别。
                   
                               #计算两个日期的间隔
                             import datetime
                             d1 = datetime.datetime.strptime(t_str1, '%Y%m')
                             d2 = datetime.datetime.strptime(t_str2, '%Y%m')
                             delta = d1 - d2
                             d3 = abs(delta.days)
                             #print(d3)
                             if b != 0 and d3 != 0:
                                   shu = (c-b)/b*(365/d3)
                                   print("展期收益")
                                   print(shu)
                                   df = OI1
                                   df = df.replace("[","")
                                   df = df.replace("]","")
                                   df = df.replace("'","")
                                   df = df.replace(" ","")
                                   OI1 = df
                                   df = OI2
                                   df = df.replace("[","")
                                   df = df.replace("]","")
                                   df = df.replace("'","")
                                   df = df.replace(" ","")
                                   OI2 = df
                                   writer2.writerow([shift_date,line2,shu,OI1,OI2])#写入标签                               
                             else:
                                   print("展期收益无法计算")
                                   print('X')
                                   #arr1[m] = 0
                             
                             shu = 0
                             c = 0
                             b = 0
                             openInterest1 = 0
                             openInterest2 = 0
                       """      
                       t2.seek(4)
                       dd = t2.read()                            
                       dd = dd.split("\r\n") #or list2 = str2.split(" ")
                       t2.seek(4)
                       t2.truncate()   #清空文件t1.truncate() 
                       
                    #  print (list2)
                       while "" in dd:
                           dd.remove("")
                    #  print (list2)
                       dd = list(map(float,dd))                       
                      #df = df.date
                       """
                       t2.close()#需要在此处分成两个程式，目标合约已经改成了远月合约。
                       """
                       df = pd.read_csv("report.csv",encoding='gb2312')
                       print(df)
                       dd = df[df['DAY'].isin([shift_date])]
                       print(dd)
                       dd = dd.DATA
                       dd = list(dd)
                       print(dd)
                       bubbleSort(dd)
                       
                       
                       arr3 =[]  
                       arr4 =[]                        
                       l = 0
                       for pp in dd:                           
                           if l <= int(0.2*(len(dd))):
#                               df = df.loc[df['S_INFO_WINDCODE'].str.contains(line2)] #筛选（包含）
#                               df.reset_index(inplace=True)
#                               pp = str(pp)
                                print(pp)
                                df = pd.read_csv("report.csv",encoding='GB2312')
                                df = df[df['DAY'].isin([shift_date])]
                                df = df[df['DATA'] == pp]
                                df = df.ID.tolist()
                                df = str(df)
                                df = df.replace("[","")
                                df = df.replace("]","")
                                df = df.replace("'","")
                                df = df.replace(" ","")
                                print(df)
                                arr3.append(df) 
                           if  l >= int(0.8*(len(dd))) and l <= len(dd):
#                               df = df.loc[df['DATA'].str.contains(str(pp))] 
#                               dd = str(pp)
                                print(pp)
                                df = pd.read_csv("report.csv",encoding='GB2312')
                                df = df[df['DAY'].isin([shift_date])]
                                df = df[df['DATA'] == pp]
                                df = df.ID.tolist() 
                                df = str(df)                                          
                                df = df.replace("[","")
                                df = df.replace("]","")
                                df = df.replace("'","")
                                df = df.replace(" ","")
                                print(df)
                                arr4.append(df) 
                           l = l+1
                       thscode1 = arr3
                       thscode2 = arr4 
                       
                       arr5 = [] 
                       arr6 = []
                       arr2 = []
                       
                       for OIO in arr3:
                           df = pd.read_csv("report.csv",encoding='GB2312')
                           df = df[df['DAY'].isin([shift_date])]
                           for IOI in df.ID:
                               if OIO == IOI:
                                   df = pd.read_csv("report.csv",encoding='GB2312')
                                   df = df[df['DAY'].isin([shift_date])]
                                   df = df[df['ID'] == IOI]
                                   df = str(df.Submaincontract.values)
                                   df = df.replace("[","")
                                   df = df.replace("]","")
                                   df = df.replace("'","")
                                   df = df.replace(" ","")
                                   arr5.append(df) 
                      #print(arr5)
                       for OIO in arr4:
                           df = pd.read_csv("report.csv",encoding='GB2312')
                           df = df[df['DAY'].isin([shift_date])]
                           for IOI in df.ID:
                               if OIO == IOI:
                                   df = pd.read_csv("report.csv",encoding='GB2312')
                                   df = df[df['DAY'].isin([shift_date])]
                                   df = df[df['ID'] == IOI]
                                   df = str(df.Submaincontract.values)
                                   df = df.replace("[","")
                                   df = df.replace("]","")
                                   df = df.replace("'","")
                                   df = df.replace(" ","")
                                   arr6.append(df)  
                      
                       arr2 = shift_date,thscode1,thscode2,arr5,arr6
                       writer3.writerows([arr2])#写入标签 
#                      df = pd.read_csv("report.csv",encoding='GB2312')
#                      df = df.drop(columns=["DAY","ID","DATA","Maincontract","Submaincontract"])
#                      df.to_csv("report.csv",index=False,encoding='GB2312')
                       """
                       
    
    

    




